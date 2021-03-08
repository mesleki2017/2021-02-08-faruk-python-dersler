from openpyxl import Workbook
import openpyxl

workbook = openpyxl.load_workbook("excel//veriyazdim.xlsx")

sheet = workbook.active

print(sheet.cell(row=1, column=1).value)

print(sheet.cell(row=2, column=1).value)

print(len(sheet[1]))

print(sheet["A"])