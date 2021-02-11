# https://zetcode.com/python/openpyxl/

import openpyxl
from openpyxl import Workbook


def listeyi_excele_satir_olarak_kaydet(liste, satir, sutun, excel_dosya_ismi):
    try:
        workbook = openpyxl.load_workbook(excel_dosya_ismi+".xlsx")
    except:
        workbook = Workbook()
    sheet = workbook.active
    for iii in range(0, len(liste)):
        sheet.cell(row=satir, column=sutun+iii).value = liste[iii]
    workbook.save(filename=excel_dosya_ismi+".xlsx")


# liste1 = [8, 944, 6, 7, 9, 2]

# liste2 = [82, 9, 699, 77, 89, 332]

# listeyi_excele_satir_olarak_kaydet(liste1, 4, 2, "datam1")

# listeyi_excele_satir_olarak_kaydet(liste2, 9, 3, "datam1")
