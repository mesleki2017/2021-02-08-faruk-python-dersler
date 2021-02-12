from openpyxl import Workbook
import numpy as np


time = np.arange(-3*np.pi, 3*np.pi, 0.01)

amplitude = np.sin(time)

workbook = Workbook()
sheet = workbook.active

sheet.cell(row=1, column=1).value = "zaman"
sheet.cell(row=1, column=2).value = "sinus"

for iii in range(0, len(time)):
    sheet.cell(row=iii+2, column=1).value = time[iii]
    sheet.cell(row=iii+2, column=2).value = amplitude[iii]


workbook.save(filename="sinus1.xlsx")
